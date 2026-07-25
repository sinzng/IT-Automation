import os
import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ───────────────────────────────────────
# 환경변수 읽기
# ───────────────────────────────────────
host_ip = os.environ.get("host_ip", "")
jeus_id = os.environ.get("jeus_id", "")
device_name = os.environ.get("device_name", "")
port = os.environ.get("port", "")
json_path = os.environ.get("json_path", "")
output_path = os.environ.get("output_path", "")
file_name = os.environ.get("file_name", "")

# ───────────────────────────────────────
# 엑셀로 저장 (openpyxl 사용)
# ───────────────────────────────────────

if not json_path or not os.path.isdir(json_path):
    print(f"JSON 경로가 존재하지 않거나 잘못되었습니다: {json_path}")
    exit(1)

os.makedirs(output_path, exist_ok=True)
xlsx_file = os.path.join(output_path, file_name)

# ───────────────────────────────────────
# 엑셀 초기화
# ───────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Report"

headers = ["장비명", "IP Address", "Hostname", "JEUS ID", "만료일", "확인자", "자동변경", "수동확인", "비고"]
ws.append(headers)

# 스타일 정의
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="666666")
center_align = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# ───────────────────────────────────────
# JSON 순회 → 엑셀로 추가
# ───────────────────────────────────────
for fname in sorted(os.listdir(json_path)):
    if not fname.endswith(".json"):
        continue

    file_path = os.path.join(json_path, fname)
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            item = json.load(f)

        auto_flag = "O"
        remark = "변경완료" if item.get("status") == "변경완료" else "변경실패"

        row = [
            item.get("device", ""),
            item.get("host", ""),
            item.get("host_name", ""),
            item.get("jeus_id", ""),  
            item.get("due_day", ""),
            "",          # 확인자
            auto_flag,
            "",          # 수동확인
            remark
        ]
        ws.append(row)

    except Exception as e:
        print(f"{fname} 파일 읽기 실패: {e}")

# ───────────────────────────────────────
# 스타일 적용
# ───────────────────────────────────────
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
    for cell in row:
        cell.alignment = center_align
        cell.border = border
        if cell.row == 1:
            cell.font = header_font
            cell.fill = header_fill

# 열 너비
column_widths = [25, 18, 18, 12, 14, 12, 10, 12, 12]
for i, width in enumerate(column_widths, start=1):
    col_letter = chr(64 + i)
    ws.column_dimensions[col_letter].width = width

# 저장
wb.save(xlsx_file)
print(f"Excel 저장 완료: {xlsx_file}")